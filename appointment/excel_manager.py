"""
Excel Manager for Appointment Data

Thread-safe CRUD operations on the appointments Excel file.
Adapted from voice_appointment_bot/excel/manager.py.
"""

import os
import time
import threading
import logging
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()
logger = logging.getLogger(__name__)

# Thread lock — multiple bot users could write simultaneously
_excel_lock = threading.Lock()

# Resolve path — use the main project's root as base
_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_env_excel = os.getenv("EXCEL_FILE_PATH", "excel/appointments.xlsx")
EXCEL_PATH = os.path.join(_BASE_DIR, _env_excel)

logger.info(f"[Appointment Excel] File location: {EXCEL_PATH}")

COLUMNS = [
    "appointment_id",
    "security_token",
    "patient_name",
    "patient_age",
    "patient_phone",
    "telegram_chat_id",
    "preferred_date",
    "preferred_time",
    "symptoms",
    "status",
    "confirmed_date",
    "confirmed_time",
    "doctor_notes",
    "created_at",
    "updated_at",
]


def _ensure_workbook_exists():
    """Creates the Excel file with headers ONLY if it does not already exist."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Appointments"

        HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        HEADER_FONT = Font(bold=True)

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(col_name) + 4)

        wb.save(EXCEL_PATH)
        logger.info(f"[Appointment Excel] Created new file: {EXCEL_PATH}")


def write_appointment(appointment: dict) -> None:
    """Appends a new appointment row. Thread-safe with retries."""
    import openpyxl

    _ensure_workbook_exists()

    max_retries = 5
    for attempt in range(max_retries):
        with _excel_lock:
            try:
                wb = openpyxl.load_workbook(EXCEL_PATH)
                ws = wb.active
                row_values = [appointment.get(col, "") for col in COLUMNS]
                ws.append(row_values)
                wb.save(EXCEL_PATH)
                logger.info(f"[Appointment Excel] Written: {appointment.get('appointment_id')}")
                return
            except PermissionError as e:
                logger.warning(f"[Appointment Excel] Locked (attempt {attempt+1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(2)
                else:
                    raise
            except Exception as e:
                logger.error(f"[Appointment Excel] Write failed: {e}", exc_info=True)
                raise


def get_appointment_by_id(appointment_id: str) -> dict | None:
    """Searches for an appointment by ID. Returns dict or None."""
    import openpyxl

    _ensure_workbook_exists()

    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == appointment_id:
                return dict(zip(COLUMNS, row))
    return None


def update_appointment_status(
    appointment_id: str,
    security_token: str,
    new_status: str,
    confirmed_date: str = "",
    confirmed_time: str = "",
    doctor_notes: str = "",
) -> dict | None:
    """Updates status after security_token validation."""
    import openpyxl

    _ensure_workbook_exists()

    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            cell_id = row[0].value
            cell_token = row[1].value

            if cell_id == appointment_id:
                if cell_token != security_token:
                    logger.warning(f"[Appointment Excel] Token mismatch for {appointment_id}")
                    return None

                now = datetime.now().isoformat(timespec="seconds")

                ws.cell(row=row_idx, column=COLUMNS.index("status") + 1).value = new_status
                ws.cell(row=row_idx, column=COLUMNS.index("confirmed_date") + 1).value = confirmed_date
                ws.cell(row=row_idx, column=COLUMNS.index("confirmed_time") + 1).value = confirmed_time
                ws.cell(row=row_idx, column=COLUMNS.index("doctor_notes") + 1).value = doctor_notes
                ws.cell(row=row_idx, column=COLUMNS.index("updated_at") + 1).value = now

                wb.save(EXCEL_PATH)
                logger.info(f"[Appointment Excel] {appointment_id} → {new_status}")

                updated_row = [cell.value for cell in ws[row_idx]]
                return dict(zip(COLUMNS, updated_row))

        logger.warning(f"[Appointment Excel] Not found: {appointment_id}")
        return None


def is_slot_taken(preferred_date: str, preferred_time: str) -> bool:
    """Returns True if there is a Pending/Confirmed appointment at the given slot."""
    import openpyxl

    _ensure_workbook_exists()

    max_retries = 5
    for attempt in range(max_retries):
        with _excel_lock:
            try:
                wb = openpyxl.load_workbook(EXCEL_PATH)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(COLUMNS, row))
                    if (
                        row_dict.get("preferred_date") == preferred_date
                        and row_dict.get("preferred_time") == preferred_time
                        and row_dict.get("status") in ("Pending", "Confirmed")
                    ):
                        return True
                return False
            except PermissionError:
                logger.warning(f"[Appointment Excel] Locked during slot check ({attempt+1}/{max_retries})")
                if attempt < max_retries - 1:
                    time.sleep(2)
                else:
                    return False
    return False
